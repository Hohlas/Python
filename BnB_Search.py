from datetime import datetime, date, time
import time

import openpyxl # pip install openpyxl
from openpyxl import Workbook
from BnB_Lib import *

KeyFile="API_KEY.txt"
with open(KeyFile, "r") as myfile:
    line = myfile.readline() # Читает строку из файла
    API_SECRET=myfile.readline() # Читает строку из файла
myfile.close()
API_KEY=line.split('\n')[0] # читает строку до символов '..'
from binance_api import Binance 
bot = Binance(API_KEY=API_KEY,  API_SECRET=API_SECRET)

TradeLot=0.0015

def TIME(view=''):
    if view=='full':
        return datetime.today().strftime('%d.%m.%Y %H:%M:%S')
    else:
        return datetime.today().strftime('%H:%M:%S')


# INFO = bot.exchangeInfo()                 # Настройки и лимиты биржи 
# INFO - bot.depth(symbol=SYM,limit=5)      # Открытые ордера (книгa ордеров), limit - кол-во возвращаемых записей от 5 до 1000 
# INFO = bot.trades(symbol=SYM,limit=1)     # Последние (чужие) сделки
# INFO = bot.klines(symbol=SYM, interval='5m', limit=1) # Данные по свечам, limit – кол-во свечей (максимум 500, по умолчанию 500)
        # 1499040000000,      // Время открытия
        # "0.01634790",       // Цена открытия (Open)
        # "0.80000000",       // Максимальная цена (High)
        # "0.01575800",       // Минимальная цена (Low)
        # "0.01577100",       // Цена закрытия (Close)
        # "148976.11427815",  // Объем
        # 1499644799999,      // Время закрытия
        # "2434.19055334",    // Объем квотируемой валюты
        # 308,                // Кол-во сделок
        # "1756.87402397",    // Taker buy base asset volume
        # "28.46694368",      // Taker buy quote asset volume
        # "17928899.62484339" // Ignore
# INFO = bot.aggTrades(symbol=SYM, limit=1) # Сжатая история сделок     {"p":"1.01", "q":"4.7", "T":1498793709153}      
# INFO = bot.tickerPrice(symbol=SYM)       # Last Price {"symbol": "LTCBTC",  "price": "4.0"} 
# INFO = bot.tickerBookTicker(symbol=SYM)  # Bid Ask    {"symbol":"LTCBTC", "bidPrice":"4.0", "bidQty":"431", "askPrice":"4.0", "askQty":"9"}
# INFO = bot.ticker24hr(symbol=SYM)        # Статистика за 24 часа  {..."bidPrice": "4.0", "askPrice": "4.0",  


def STAT24(LIST):
    DATA = bot.ticker24hr()   # Best prices { "symbol": "LTCBTC",  "bidPrice": x,  "bidQty": x,  "askPrice": x,  "askQty": x }
    for pair in LIST: #  range(0, 5, 1)    
        for Ticker in DATA: # find in DATA string with the same Symbol 
            if pair['Sym']==Ticker['symbol']: break 
        pair.update({'Time':    str(datetime.now().strftime('%H:%M:%S'))})
        pair.update({'Bid':     float(Ticker['bidPrice'])})  # current value
        pair.update({'Ask':     float(Ticker['askPrice'])})  # current value
        pair.update({'Vol':     int(float(Ticker['quoteVolume']))})
        pair.update({'Trades':  float(Ticker['count'])})
        pair.update({'Spred':   pair['Ask']-pair['Bid']})
        if pair['Bid']>0:       pair.update({'Rank':    round(pair['Spred']/pair['Bid']*100, 2)  })
        else:                   pair.update({'Rank': 0})

def BIDASK(LIST):
    DATA = bot.tickerBookTicker()   # Best prices { "symbol": "LTCBTC",  "bidPrice": x,  "bidQty": x,  "askPrice": x,  "askQty": x }
    for pair in LIST: #  range(0, 5, 1)    
        for Ticker in DATA: # find in DATA string with the same Symbol 
            if pair['Sym']==Ticker['symbol']: break 
        pair.update({'Time':    str(datetime.now().strftime('%H:%M:%S'))})# update time
        pair.update({'Bid':     float(Ticker['bidPrice'])})
        pair.update({'Ask':     float(Ticker['askPrice'])})
        pair.update({'BidQty':  float(Ticker['bidQty'])})
        pair.update({'AskQty':  float(Ticker['askQty'])}) 
        pair.update({'Spred':   pair['Ask']-pair['Bid']})
        if pair['Bid']>0:       pair.update({'Rank':    pair['Spred']/pair['Bid']})
        else:                   pair.update({'Rank':0})  
    
def CANDLE_INFO(LIST, per='5m'): # very slow function because of many requests       
    for pair in LIST: #  range(0, 5, 1)    
        Candle = bot.klines(symbol=pair['Sym'], interval=per, limit=2) # get list of last two candles [[1532604240000, '0.05807400', '0.05810900',.., '7.74028777', '0']]
        ot,o,h,l,c,v,ct,vq,n,bb,bq,i = Candle[0] # Candle[1] don't complete 
        pair.update({'High':    float(h)})
        pair.update({'Low':     float(l)})
        pair.update({'Close':   float(c)})
        pair.update({'Vol':     float(vq)})
        pair.update({'Trades':  float(n)})

def AVG_VOL_FOR_LAST_N_TRDS(LIST): # very slow function because of many requests   
    N=10  # last trades quantity
    price=0; quan=0; 
    for pair in LIST: #     
        TRADES = bot.aggTrades(symbol=pair['Sym'], limit=N)     # last N trades of ALL traders {"p":"1.01", "q":"4.7", "T":1498793709153}
        for trade in TRADES:
            price+=float(trade['p']) 
            quan+=float(trade['q'])  
        pair.update({'price': price/N}) # average price and 
        pair.update({'quan':  quan/N})  # quantity for last N trades
        pair.update({'Freq': int(TRADES[N-1]['T']-TRADES[0]['T'])}) # trades frequency = time between first and last trades


def FULL_LIST_SCAN():
    ExchangeInfo = bot.exchangeInfo() # All pairs list and settings
    LIST=[] # create empty list of all BTC pairs
    for Inf in ExchangeInfo['symbols']: # in all cells in ExchangeInfo 
        if Inf['quoteAsset']!='BTC': continue # need only symbols with BTC quoteAsset 
        LIST.append({
            'Sym':                  Inf['symbol'], 
            'baseAsset':            Inf['baseAsset'], 
            'quoteAsset':           Inf['quoteAsset'], 
            'baseAssetPrecision':   Inf['baseAssetPrecision'], 
            'filters':              Inf['filters'],
            'lot':                  0,           # add 
            'Ord':                  'NONE',                 # necessary 
            'BaseBal':              {'Free':0,'Lock':0},# cells 
            'QuoteBal':             {'Free':0,'Lock':0}
            })
    local_time = int(time.time())
    server_time = int(ExchangeInfo['serverTime'])//1000
    shift_seconds = server_time-local_time
    bot.set_shift_seconds(shift_seconds) # set shift_seconds for trading
    print(TIME('full'), 'Start program. Time difference =',shift_seconds,'sec')
    return LIST
        
def SORT_LIST(LIST, SortedBy='Rank', LimVal=1):
    SORTED_LIST=sorted(LIST, key= lambda d: d[SortedBy], reverse=True)   # сортировка списка словарей по убыванию 'RANK'
    n=-1
    for Pair in SORTED_LIST:            # in all the cells
        n+=1                            # remember cell number
        if Pair[SortedBy]<LimVal: break # until value 'SortedBy' is correct 
    return SORTED_LIST[0: n] # get only satisfactory cells


def BEST_PAIRS(FULL_LIST):    
    STAT24(FULL_LIST)   # update FULL_LIST with 24H parameters: Bid,Ask,Vol,Trades,Spred,Rank 
    SORTED1 = SORT_LIST(FULL_LIST,SortedBy='Rank',    LimVal=PROFIT) # get Part of FULL_LIST with Rank better than PROFIT
    SORTED2 = SORT_LIST(SORTED1,  SortedBy='Trades',  LimVal=0) # sort SORTED1 by 'Trades per 24H'
    SORTED2 = SORTED2[0: int(len(SORTED2)/2)] # get better half of SORTED2 to speed up further search
    CANDLE_INFO(SORTED2,'5m') # get parameters (trades per 5M) of 5M candle for SORTED2. SORTED2 list must be small, because of CANDLE_INFO function too slow.  
    BEST_LIST  =SORT_LIST(SORTED2, SortedBy='Trades', LimVal=0) # sort SORTED2 by 'Trades per 5M'
    BEST_LIST = BEST_LIST[0: int(len(BEST_LIST)/2)] # get better half BEST_LIST
    SYM_LIST=list()
    for Pair in BEST_LIST:
        Pair['lot']=TradeLot
        SYM_LIST.append(Pair['Sym']) # create list with used symbols
    # add symbols with open orders from last trades
    LimOrders=bot.openOrders()
    for Ord in LimOrders: # in all current limit orders    
        if Ord['symbol'] not in SYM_LIST: 
            for Pair in FULL_LIST: 
                if Pair['Sym']==Ord['symbol']: break # find necessary Sym in FULL_LIST      
            Pair['Ord']=Ord # update order info
            BEST_LIST.append(Pair)     # insert new sym 
            SYM_LIST.append(Pair['Sym'])# insert new sym 
        else: Pair.update({'Ord':  Ord})# update sym info
    # search all currences in depo (from last nonclosed trades) to sell them with profit
    Balances=bot.account()['balances']
    for Bal in Balances:
        if float(Bal['free'])<0.0001: continue # skip empty balances
        if Bal['asset']=='BTC': # there is no BTCBTC pair. 
            BALANCE=float(Bal['free'])+float(Bal['locked'])
            continue
        if Bal['asset']=='BNB': continue # never sell BNB, it is needed for pay comission
        Sym=Bal['asset']+'BTC'
        if Sym not in SYM_LIST:  # 
            SYM_LIST.append(Sym) # update list with used symbols 
            for Pair in FULL_LIST:
                if Sym==Pair['Sym']:   
                    BEST_LIST.append(Pair)
                    BALANCE+=(float(Bal['free'])+float(Bal['locked']))*Pair['Ask']
    BEST_LIST=sorted(BEST_LIST, key= lambda d: d['lot'], reverse=False)   # Pairs must be closed in top of list
    print('BALANCE=',BALANCE)
    return BEST_LIST, SYM_LIST

def BALANCE_UPDATE(PAIRS):  ##################################################################################################################################################
    Balances=bot.account()['balances']
    for Bal in Balances:
        for Pair in PAIRS:
            if Bal['asset']==Pair['baseAsset']:  Pair.update({'BaseBal':   {'Free':float(Bal['free']), 'Lock':float(Bal['locked'])}}) # update dict Pair['Bal']['Base']
            if Bal['asset']==Pair['quoteAsset']: Pair.update({'QuoteBal':  {'Free':float(Bal['free']), 'Lock':float(Bal['locked'])}}) # update dict Pair['Bal']['Quote']
    


def XLS_SAVE(LIST): ### Save received lists to Exel sheets  ######################################################################################################
    BALANCE_UPDATE(LIST)
    XlsFile=str(datetime.today().strftime('BnB_%m%d.xlsx'))
    try:
        book = openpyxl.load_workbook(XlsFile) # try to Load XLSX file  
    except Exception  as excpt:# общее исключение, под которое попадают все исключительные ситуации
        book = Workbook()
    sheet = book.create_sheet(str(datetime.now().strftime('PAIRS_%H_%M')), 0)
    sheet.append(['Sym','Bid','Ask','Vol','Trades','Spred','Rank','lot','BTC_Bal','Ord'])
    for Pair in LIST:
        BTC_Bal=max(float(Pair['BaseBal']['Free']), float(Pair['BaseBal']['Lock'])) * float(Pair['Ask'])
        if Pair['Ord']=='NONE': Order='NONE'
        else: Order=Pair['Ord']['status']+Pair['Ord']['side']
        sheet.append([Pair['Sym'],Pair['Bid'],Pair['Ask'],Pair['Vol'],Pair['Trades'],Pair['Spred'],Pair['Rank'],Pair['lot'],BTC_Bal,Order])
    try:
        book.save(XlsFile)
    except Exception  as excpt:  # file is busy
        XlsFile=str(datetime.today().strftime('BnB_%m%d_copy.xlsx')) # create another name
        book.save(XlsFile)
    print("Create file ",XlsFile)

   
 
