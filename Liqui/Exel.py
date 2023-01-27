from openpyxl import Workbook
wb = Workbook()
ws = wb.active  # Выбрать активную книгу
ws['A1'] = 42  # Запишем число 42 в ячейку А1
ws.append([1, 2, 3])  # Добавим строку с данными

import datetime
ws['A2'] = datetime.datetime.now()# В поле А2 запишем текущую дату и время

import os
wb.save(os.path.dirname(os.path.abspath(__file__)) + "/sample.xlsx")# Сохраним файл рядом со скриптом