from datetime import datetime, date, time
import time

import openpyxl # pip install openpyxl
from openpyxl import Workbook
from BnB_Lib import *



def TIME(view=''):
    if view=='full':
        return datetime.today().strftime('%d.%m.%Y %H:%M:%S')
    else:
        return datetime.today().strftime('%H:%M:%S')

def UTIME(utime): # convert UNIX time to 'MM.DD HH.MM.SS' format
    t=time.gmtime(utime)    # GM format
    # t=time.localtime(utime) # local time format
    return time.strftime('%m.%d %H:%M:%S',t)


import urllib.request
def TGM_SEND(text): # send Telegram Missage
    urllib.request.urlopen("""
        https://api.telegram.org/bot{API_TOKEN}/sendMessage?chat_id={CHAT_ID}&text={TEXT}
    """.format(
        API_TOKEN = '686961167:AAFqI7OZ3scrm8Qp-XxO0sN_242aHBnzCmM',
        CHAT_ID = '332155342',
        TEXT = text
    ))
    print(TIME(),text)

def TGM_GET(): # send Telegram Missage
    urllib.request.urlopen("""
        https://api.telegram.org/bot{API_TOKEN}/getUpdates
    """.format(
        API_TOKEN = '686961167:AAFqI7OZ3scrm8Qp-XxO0sN_242aHBnzCmM',
        CHAT_ID = '332155342',
        TEXT = text
    ))



TOKEN = '686961167:AAFqI7OZ3scrm8Qp-XxO0sN_242aHBnzCmM'
import telebot   # pip3.6 install --user pyTelegramBotAPI
tm = telebot.TeleBot(TOKEN)

 

def XLS_SAVE(LIST): ### Save received lists to Exel sheets  ######################################################################################################
    
    XlsFile=str(datetime.today().strftime('BnB_%m%d.xlsx'))
    try:
        book = openpyxl.load_workbook(XlsFile) # try to Load XLSX file  
    except Exception  as excpt:# общее исключение, под которое попадают все исключительные ситуации
        book = Workbook()
    sheet = book.create_sheet(str(datetime.now().strftime('%H_%M')), 0)
    sheet.append(['Sym','Bid','Ask','Vol','Trades','Spred','BidAsk','lot','BTC_Bal','Ord'])
    for Pair in LIST:
        BTC_Bal=(float(Pair['BaseBal']['Free'])+ float(Pair['BaseBal']['Lock'])) * float(Pair['Ask'])
        if Pair['Ord']=='NONE': Order='NONE'
        else: Order=Pair['Ord']['status']+Pair['Ord']['side']
        sheet.append([Pair['Sym'],Pair['Bid'],Pair['Ask'],Pair['Vol'],Pair['Trades'],Pair['Spred'],Pair['BidAsk'],Pair['lot'],BTC_Bal,Order])
    try:
        book.save(XlsFile)
    except Exception  as excpt:  # file is busy
        XlsFile=str(datetime.today().strftime('BnB_%m%d_copy.xlsx')) # create another name
        book.save(XlsFile)
    print("Create file ",XlsFile)

   
 
