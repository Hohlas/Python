from datetime import datetime, date, time
import time

import openpyxl # pip install openpyxl
from openpyxl import Workbook



KeyFile="API_KEY.txt"
with open(KeyFile, "r") as myfile:
    line = myfile.readline() # Читает строку из файла
    API_SECRET=myfile.readline() # Читает строку из файла
myfile.close()
API_KEY=line.split('\n')[0] # читает строку до символов '..'

from binance_api import Binance 
bot = Binance(
    API_KEY=API_KEY,
    API_SECRET=API_SECRET
    )
print(datetime.today().strftime('%d.%m.%Y %H:%M:%S'))

def AAA(LIST):
   LIST.append("John")



# INFO = bot.exchangeInfo()                 # Настройки и лимиты биржи 
# INFO - bot.depth(symbol=SYM,limit=5)      # Открытые ордера (книгa ордеров), limit - кол-во возвращаемых записей от 5 до 1000 
# INFO = bot.trades(symbol=SYM,limit=1)     # Последние (чужие) сделки
# INFO = bot.klines(symbol=SYM, interval='5m', limit=1) # Данные по свечам, limit – кол-во свечей (максимум 500, по умолчанию 500)
        # 1499040000000,      // Время открытия
        # "0.01634790",       // Цена открытия (Open)
        # "0.80000000",       // Максимальная цена (High)
        # "0.01575800",       // Минимальная цена (Low)
        # "0.01577100",       // Цена закрытия (Close)
        # "148976.11427815",  // Объем
        # 1499644799999,      // Время закрытия
        # "2434.19055334",    // Объем квотируемой валюты
        # 308,                // Кол-во сделок
        # "1756.87402397",    // Taker buy base asset volume
        # "28.46694368",      // Taker buy quote asset volume
        # "17928899.62484339" // Ignore
# INFO = bot.aggTrades(symbol=SYM, limit=1) # Сжатая история сделок     {"p":"1.01", "q":"4.7", "T":1498793709153}      
# INFO = bot.tickerPrice(symbol=SYM)       # Last Price {"symbol": "LTCBTC",  "price": "4.0"} 
# INFO = bot.tickerBookTicker(symbol=SYM)  # Bid Ask    {"symbol":"LTCBTC", "bidPrice":"4.0", "bidQty":"431", "askPrice":"4.0", "askQty":"9"}
# INFO = bot.ticker24hr(symbol=SYM)        # Статистика за 24 часа  {..."bidPrice": "4.0", "askPrice": "4.0",  


def STAT24(LIST):
    DATA = bot.ticker24hr()   # Best prices { "symbol": "LTCBTC",  "bidPrice": x,  "bidQty": x,  "askPrice": x,  "askQty": x }
    for pair in LIST: #  range(0, 5, 1)    
        for Ticker in DATA: # find in DATA string with the same Symbol 
            if pair['Sym']==Ticker['symbol']: break 
        pair.update({'Time':    str(datetime.now().strftime('%H:%M:%S'))})
        pair.update({'Bid':     float(Ticker['bidPrice'])})  # current value
        pair.update({'Ask':     float(Ticker['askPrice'])})  # current value
        pair.update({'Vol':     int(float(Ticker['quoteVolume']))})
        pair.update({'Trades':  float(Ticker['count'])})
        pair.update({'Spred':   pair['Ask']-pair['Bid']})
        if pair['Bid']>0:       pair.update({'Rank':    pair['Spred']/pair['Bid']})
        else:                   pair.update({'Rank': 0})

def BIDASK(LIST):
    DATA = bot.tickerBookTicker()   # Best prices { "symbol": "LTCBTC",  "bidPrice": x,  "bidQty": x,  "askPrice": x,  "askQty": x }
    for pair in LIST: #  range(0, 5, 1)    
        for Ticker in DATA: # find in DATA string with the same Symbol 
            if pair['Sym']==Ticker['symbol']: break 
        pair.update({'Time':    str(datetime.now().strftime('%H:%M:%S'))})# update time
        pair.update({'Bid':     float(Ticker['bidPrice'])})
        pair.update({'Ask':     float(Ticker['askPrice'])})
        pair.update({'BidQty':  float(Ticker['bidQty'])})
        pair.update({'AskQty':  float(Ticker['askQty'])}) 
        pair.update({'Spred':   pair['Ask']-pair['Bid']})
        if pair['Bid']>0:       pair.update({'Rank':    pair['Spred']/pair['Bid']})
        else:                   pair.update({'Rank':0})  
    
def CANDLE_INFO(LIST, per='5m'): # very slow function because of many requests       
    for pair in LIST: #  range(0, 5, 1)    
        Candle = bot.klines(symbol=pair['Sym'], interval=per, limit=2) # get list of last two candles [[1532604240000, '0.05807400', '0.05810900',.., '7.74028777', '0']]
        ot,o,h,l,c,v,ct,vq,n,bb,bq,i = Candle[0] # Candle[1] don't complete 
        pair.update({'High':    float(h)})
        pair.update({'Low':     float(l)})
        pair.update({'Close':   float(c)})
        pair.update({'Vol':     float(vq)})
        pair.update({'Trades':  float(n)})

def AVG_VOL_FOR_LAST_N_TRDS(LIST): # very slow function because of many requests   
    N=10  # last trades quantity
    price=0; quan=0; 
    for pair in LIST: #     
        TRADES = bot.aggTrades(symbol=pair['Sym'], limit=N)     # last N trades of ALL traders {"p":"1.01", "q":"4.7", "T":1498793709153}
        for trade in TRADES:
            price+=float(trade['p']) 
            quan+=float(trade['q'])  
        pair.update({'price': price/N}) # average price and 
        pair.update({'quan':  quan/N})  # quantity for last N trades
        pair.update({'Freq': int(TRADES[N-1]['T']-TRADES[0]['T'])}) # trades frequency = time between first and last trades


def FULL_LIST_SCAN():
    ExchangeInfo = bot.exchangeInfo() # All pairs list and settings
    LIST=[] # create empty list of all BTC pairs
    for Inf in ExchangeInfo['symbols']: # in all cells in ExchangeInfo 
        if Inf['quoteAsset']!='BTC': continue # need only symbols with BTC quoteAsset 
        LIST.append({'Sym':Inf['symbol'], 'baseAsset':Inf['baseAsset'], 'quoteAsset':Inf['quoteAsset'], 'filters':Inf['filters']})
    return LIST
        
def SORT_LIST(ORIGINAL_LIST, SortedBy='Rank', Part=3):
    SORTED_LIST=sorted(ORIGINAL_LIST, key= lambda d: d[SortedBy], reverse=True)   # сортировка списка словарей по убыванию 'RANK'
    ROWS=int(len(ORIGINAL_LIST)/Part)
    return SORTED_LIST[0:ROWS]

def BEST_PAIRS(FULL_LIST):    
    INIT=FULL_LIST
    STAT24(FULL_LIST)    # update FULL_LIST with 24H parameters: Bid,Ask,Vol,Trades,Spred,Rank 
    BEST_RANK   =SORT_LIST(FULL_LIST,   SortedBy='Rank',    Part=3)
    BEST_TRADES =SORT_LIST(BEST_RANK,   SortedBy='Trades',  Part=3)
    CANDLE_INFO(BEST_TRADES,'5m')
    BEST_PAIRS  =SORT_LIST(BEST_TRADES, SortedBy='Rank',    Part=3)
    sheet0 = book.create_sheet(str(datetime.now().strftime('FULL_%H_%M')), 0)
    sheet1 = book.create_sheet(str(datetime.now().strftime('Trades_%H_%M')), 0)
    sheet2 = book.create_sheet(str(datetime.now().strftime('PAIRS_%H_%M')), 0)
    sheet0.append(['Sym','Bid','Ask','Vol','Trades','Spred','Rank'])
    sheet1.append(['Sym','Bid','Ask','Vol','Trades','Spred','Rank'])
    sheet2.append(['Sym','Bid','Ask','Vol','Trades','Spred','Rank'])
    for pair in FULL_LIST:      sheet0.append([pair['Sym'],pair['Bid'],pair['Ask'],pair['Vol'],pair['Trades'],pair['Spred'],pair['Rank']])
    for pair in BEST_TRADES:    sheet1.append([pair['Sym'],pair['Bid'],pair['Ask'],pair['Vol'],pair['Trades'],pair['Spred'],pair['Rank']])
    for pair in BEST_PAIRS:     sheet2.append([pair['Sym'],pair['Bid'],pair['Ask'],pair['Vol'],pair['Trades'],pair['Spred'],pair['Rank']])
    book.save(XlsFile)
    print('file saved')
    return BEST_PAIRS

XlsFile=str(datetime.today().strftime('BnB_%m%d.xlsx'))
try:
    book = openpyxl.load_workbook(XlsFile) # try to Load XLSX file
except Exception  as excpt:# общее исключение, под которое попадают все исключительные ситуации
    print("Create file ",XlsFile)
    book = Workbook()
BnB_LIST=FULL_LIST_SCAN()
PAIRS=BEST_PAIRS(BnB_LIST)
print ('ok')
        


   

   
 
